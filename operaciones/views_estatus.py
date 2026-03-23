from __future__ import annotations

from io import BytesIO
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db import models, transaction
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bitacora.models import Bitacora
from taller.models import Conductor

from .forms import EstatusOperacionalViajeForm
from .models import EstatusOperacionalViaje, TurnoEstatus


# ============================================================
# HELPERS GENERALES DEL PANEL
# ============================================================

def _parse_fecha(value: str | None):
    """
    Convierte una fecha en formato YYYY-MM-DD a date.
    Si viene vacía o inválida, devuelve la fecha local de hoy.
    """
    if not value:
        return timezone.localdate()
    try:
        return timezone.datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return timezone.localdate()


def _turnos_a_mostrar(turno_param: str):
    """
    Normaliza el parámetro de turno.
    Si viene AM o PM, devuelve solo ese turno.
    Si viene vacío o distinto, devuelve ambos.
    """
    turno_param = (turno_param or "TODOS").upper()
    if turno_param in [TurnoEstatus.AM, TurnoEstatus.PM]:
        return [turno_param]
    return [TurnoEstatus.AM, TurnoEstatus.PM]


def _get_conductores_queryset():
    """
    Obtiene conductores activos si el modelo tiene campo 'activo'.
    Si no existe ese campo, devuelve todos.
    """
    try:
        Conductor._meta.get_field("activo")
        return Conductor.objects.filter(activo=True)
    except Exception:
        return Conductor.objects.all()


def _qs_estatus(fecha, turnos, q=""):
    """
    Query base del panel de estatus de viajes.
    Incluye relaciones frecuentes y filtro general de búsqueda.
    """
    qs = (
        EstatusOperacionalViaje.objects
        .select_related("conductor", "tracto", "rampla", "cliente")
        .filter(fecha=fecha, turno__in=turnos)
        .order_by("conductor__apellidos", "conductor__nombres", "turno")
    )

    q = (q or "").strip()
    if q:
        qs = qs.filter(
            models.Q(conductor__nombres__icontains=q)
            | models.Q(conductor__apellidos__icontains=q)
            | models.Q(conductor__rut__icontains=q)
            | models.Q(tracto__patente__icontains=q)
            | models.Q(rampla__patente__icontains=q)
            | models.Q(cliente__razon_social__icontains=q)
            | models.Q(cliente__rut__icontains=q)
            | models.Q(nro_guia__icontains=q)
            | models.Q(estado_guia__icontains=q)
            | models.Q(estado_carga__icontains=q)
            | models.Q(lugar_carga__icontains=q)
            | models.Q(lugar_descarga__icontains=q)
            | models.Q(estado_texto__icontains=q)
            | models.Q(observaciones__icontains=q)
        )
    return qs


def _missing_por_turno(fecha):
    """
    Retorna dos listas:
    - conductores sin estatus AM
    - conductores sin estatus PM
    """
    conductores = _get_conductores_queryset().order_by("apellidos", "nombres")

    existentes_am = set(
        EstatusOperacionalViaje.objects.filter(fecha=fecha, turno=TurnoEstatus.AM)
        .values_list("conductor_id", flat=True)
    )
    existentes_pm = set(
        EstatusOperacionalViaje.objects.filter(fecha=fecha, turno=TurnoEstatus.PM)
        .values_list("conductor_id", flat=True)
    )

    missing_am = [c for c in conductores if c.id not in existentes_am]
    missing_pm = [c for c in conductores if c.id not in existentes_pm]
    return missing_am, missing_pm


def _estado_badge_count(qs, estado_carga):
    """
    Cuenta cuántos registros hay por tipo de estado de carga.
    """
    return qs.filter(estado_carga=estado_carga).count()


def _estatus_redirect(request: HttpRequest) -> str:
    """
    Reconstruye la URL del panel manteniendo filtros.
    Se usa al guardar/eliminar/copiar para volver al contexto actual.
    """
    fecha = request.POST.get("fecha") or request.GET.get("fecha") or ""
    turno = request.POST.get("turno") or request.GET.get("turno") or "TODOS"
    q = request.POST.get("q") or request.GET.get("q") or ""

    params = {}
    if fecha:
        params["fecha"] = fecha
    if turno:
        params["turno"] = turno
    if q:
        params["q"] = q

    url = reverse("operaciones:estatus_viajes_panel")
    return f"{url}?{urlencode(params)}" if params else url


# ============================================================
# BITÁCORA PRO
# ============================================================

def _sync_bitacora_from_estatus(bitacora: Bitacora, estatus: EstatusOperacionalViaje, request_user=None, preserve_manual=True):
    """
    Sincroniza una Bitácora desde un EstatusOperacionalViaje.

    Campos operativos que SIEMPRE se sincronizan:
    - cliente
    - conductor
    - tracto
    - rampla
    - fecha
    - fecha_arribo (desde fecha_carga)
    - fecha_descarga
    - guias_raw
    - origen
    - destino
    - descripcion_trabajo

    Campos manuales que NO se pisan si preserve_manual=True:
    - coordinador (si ya existe)
    - tarifa_flete
    - estadia
    - oc_edp_raw
    - estado

    Nota:
    - fecha_carga del estatus se mapea a fecha_arribo en bitácora
      porque así está modelado hoy tu flujo.
    """
    descripcion = (estatus.estado_texto or "").strip()
    if estatus.observaciones:
        descripcion = (
            f"{descripcion}\n{estatus.observaciones}".strip()
            if descripcion
            else estatus.observaciones.strip()
        )

    # -----------------------------
    # Datos operativos base
    # -----------------------------
    bitacora.estatus_origen = estatus
    bitacora.cliente = estatus.cliente
    bitacora.conductor = estatus.conductor
    bitacora.tracto = estatus.tracto
    bitacora.rampla = estatus.rampla

    bitacora.fecha = estatus.fecha
    bitacora.fecha_arribo = estatus.fecha_carga
    bitacora.fecha_descarga = estatus.fecha_descarga

    bitacora.origen = (estatus.lugar_carga or "").strip()
    bitacora.destino = (estatus.lugar_descarga or "").strip()
    bitacora.guias_raw = (estatus.nro_guia or "").strip()
    bitacora.descripcion_trabajo = descripcion

    # -----------------------------
    # Coordinador:
    # solo lo completamos si no existe
    # -----------------------------
    if preserve_manual:
        if not bitacora.coordinador and request_user:
            bitacora.coordinador = (
                getattr(request_user, "get_full_name", lambda: "")() or getattr(request_user, "username", "")
            )
    else:
        if request_user:
            bitacora.coordinador = (
                getattr(request_user, "get_full_name", lambda: "")() or getattr(request_user, "username", "")
            )

    return bitacora


# ============================================================
# VISTAS PRINCIPALES
# ============================================================

@login_required
@permission_required("operaciones.puede_ver_estatus_viajes", raise_exception=True)
def estatus_viajes_panel(request: HttpRequest):
    fecha = _parse_fecha(request.GET.get("fecha"))
    turno = (request.GET.get("turno") or "TODOS").upper()
    q = (request.GET.get("q") or "").strip()
    turnos = _turnos_a_mostrar(turno)

    qs = _qs_estatus(fecha, turnos, q)
    missing_am, missing_pm = _missing_por_turno(fecha)

    form = EstatusOperacionalViajeForm(
        initial={
            "fecha": fecha,
            "turno": TurnoEstatus.AM if turno == "TODOS" else turno,
        }
    )

    context = {
        "fecha": fecha,
        "turno": turno,
        "q": q,
        "items": qs,
        "form": form,
        "turnos": ["TODOS", TurnoEstatus.AM, TurnoEstatus.PM],
        "missing_am": missing_am,
        "missing_pm": missing_pm,
        "total_registros": qs.count(),
        "total_descargado": _estado_badge_count(qs, EstatusOperacionalViaje.EstadoCargaChoices.DESCARGADO),
        "total_camino_descargar": _estado_badge_count(qs, EstatusOperacionalViaje.EstadoCargaChoices.CAMINO_DESCARGAR),
        "total_retorno_vacio": _estado_badge_count(qs, EstatusOperacionalViaje.EstadoCargaChoices.RETORNO_VACIO),
    }
    return render(request, "operaciones/estatus_viajes.html", context)


@login_required
@permission_required("operaciones.puede_editar_estatus_viajes", raise_exception=True)
@require_POST
def estatus_viajes_guardar(request: HttpRequest):
    form = EstatusOperacionalViajeForm(request.POST)

    if not form.is_valid():
        messages.error(request, "Revisa el formulario: hay campos inválidos.")
        return redirect(_estatus_redirect(request))

    d = form.cleaned_data

    with transaction.atomic():
        obj, created = EstatusOperacionalViaje.objects.select_for_update().get_or_create(
            fecha=d["fecha"],
            turno=d["turno"],
            conductor=d["conductor"],
            defaults={
                "tracto": d.get("tracto"),
                "rampla": d.get("rampla"),
                "cliente": d.get("cliente"),
                "nro_guia": d.get("nro_guia"),
                "estado_guia": d.get("estado_guia"),
                "estado_carga": d.get("estado_carga"),
                "lugar_carga": d.get("lugar_carga"),
                "fecha_carga": d.get("fecha_carga"),
                "lugar_descarga": d.get("lugar_descarga"),
                "fecha_descarga": d.get("fecha_descarga"),
                "estado_texto": d.get("estado_texto"),
                "observaciones": d.get("observaciones"),
                "creado_por": request.user,
                "actualizado_por": request.user,
            },
        )

        if not created:
            # Actualización completa del registro existente
            obj.tracto = d.get("tracto")
            obj.rampla = d.get("rampla")
            obj.cliente = d.get("cliente")
            obj.nro_guia = d.get("nro_guia")
            obj.estado_guia = d.get("estado_guia")
            obj.estado_carga = d.get("estado_carga")
            obj.lugar_carga = d.get("lugar_carga")
            obj.fecha_carga = d.get("fecha_carga")
            obj.lugar_descarga = d.get("lugar_descarga")
            obj.fecha_descarga = d.get("fecha_descarga")
            obj.estado_texto = d.get("estado_texto")
            obj.observaciones = d.get("observaciones")
            obj.actualizado_por = request.user
            obj.save()

    messages.success(request, "Estatus guardado correctamente.")
    return redirect(_estatus_redirect(request))


@login_required
@permission_required("operaciones.puede_editar_estatus_viajes", raise_exception=True)
@require_POST
def estatus_viajes_eliminar(request: HttpRequest, pk: int):
    obj = get_object_or_404(EstatusOperacionalViaje, pk=pk)
    obj.delete()
    messages.success(request, "Estatus eliminado.")
    return redirect(_estatus_redirect(request))


@login_required
@permission_required("operaciones.puede_editar_estatus_viajes", raise_exception=True)
@require_POST
def estatus_viajes_copiar_am_a_pm(request: HttpRequest):
    fecha = _parse_fecha(request.POST.get("fecha"))

    am_rows = list(
        EstatusOperacionalViaje.objects.filter(fecha=fecha, turno=TurnoEstatus.AM)
    )
    creados = 0
    actualizados = 0

    for r in am_rows:
        obj, created = EstatusOperacionalViaje.objects.get_or_create(
            fecha=fecha,
            turno=TurnoEstatus.PM,
            conductor=r.conductor,
            defaults={
                "tracto": r.tracto,
                "rampla": r.rampla,
                "cliente": r.cliente,
                "nro_guia": r.nro_guia,
                "estado_guia": r.estado_guia,
                "estado_carga": r.estado_carga,
                "lugar_carga": r.lugar_carga,
                "fecha_carga": r.fecha_carga,
                "lugar_descarga": r.lugar_descarga,
                "fecha_descarga": r.fecha_descarga,
                "estado_texto": r.estado_texto,
                "observaciones": r.observaciones,
                "creado_por": request.user,
                "actualizado_por": request.user,
            },
        )

        if created:
            creados += 1
        else:
            # Copia "inteligente": solo completa campos vacíos del PM
            obj.tracto = obj.tracto or r.tracto
            obj.rampla = obj.rampla or r.rampla
            obj.cliente = obj.cliente or r.cliente
            obj.nro_guia = obj.nro_guia or r.nro_guia
            obj.estado_guia = obj.estado_guia or r.estado_guia
            obj.estado_carga = obj.estado_carga or r.estado_carga
            obj.lugar_carga = obj.lugar_carga or r.lugar_carga
            obj.fecha_carga = obj.fecha_carga or r.fecha_carga
            obj.lugar_descarga = obj.lugar_descarga or r.lugar_descarga
            obj.fecha_descarga = obj.fecha_descarga or r.fecha_descarga
            obj.estado_texto = obj.estado_texto or r.estado_texto
            obj.observaciones = obj.observaciones or r.observaciones
            obj.actualizado_por = request.user
            obj.save()
            actualizados += 1

    messages.success(
        request,
        f"Copiado AM → PM. Nuevos: {creados} | Actualizados: {actualizados}"
    )
    return redirect(
        reverse("operaciones:estatus_viajes_panel") + f"?fecha={fecha.isoformat()}"
    )


@login_required
@permission_required("operaciones.puede_ver_estatus_viajes", raise_exception=True)
@require_GET
def estatus_viajes_export_xlsx(request: HttpRequest):
    fecha = _parse_fecha(request.GET.get("fecha"))
    turno = (request.GET.get("turno") or "TODOS").upper()
    q = (request.GET.get("q") or "").strip()
    turnos = _turnos_a_mostrar(turno)

    qs = _qs_estatus(fecha, turnos, q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Estatus Viajes"

    headers = [
        "Turno",
        "Chofer",
        "RUT",
        "Pte Tracto",
        "Pte Semirremolque",
        "Estado Carga",
        "Nro Guía",
        "Estado Guía",
        "Cliente",
        "Lugar Carga",
        "Fecha Carga",
        "Lugar Descarga",
        "Fecha Descarga",
        "Observaciones",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F1F1F")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fill_ok = PatternFill("solid", fgColor="9ACD32")
    fill_warn = PatternFill("solid", fgColor="FFD966")
    fill_info = PatternFill("solid", fgColor="FFF200")

    for r in qs:
        row = [
            r.turno,
            f"{r.conductor.nombres} {r.conductor.apellidos}".strip(),
            r.conductor.rut,
            getattr(r.tracto, "patente", "") or "",
            getattr(r.rampla, "patente", "") or "",
            r.get_estado_carga_display(),
            r.nro_guia,
            r.estado_guia,
            getattr(r.cliente, "razon_social", "") or "",
            r.lugar_carga,
            r.fecha_carga.strftime("%d.%m.%Y") if r.fecha_carga else "",
            r.lugar_descarga,
            r.fecha_descarga.strftime("%d.%m.%Y") if r.fecha_descarga else "",
            r.observaciones,
        ]
        ws.append(row)

        excel_row = ws.max_row

        if r.estado_carga == EstatusOperacionalViaje.EstadoCargaChoices.DESCARGADO:
            fill = fill_ok
        elif r.estado_carga == EstatusOperacionalViaje.EstadoCargaChoices.CAMINO_DESCARGAR:
            fill = fill_info
        else:
            fill = fill_warn

        for cell in ws[excel_row]:
            cell.fill = fill

    widths = {
        "A": 8,
        "B": 28,
        "C": 16,
        "D": 14,
        "E": 18,
        "F": 20,
        "G": 14,
        "H": 16,
        "I": 24,
        "J": 22,
        "K": 14,
        "L": 22,
        "M": 14,
        "N": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"estatus_viajes_{fecha.strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
@permission_required("operaciones.puede_editar_estatus_viajes", raise_exception=True)
@require_GET
def estatus_viajes_a_bitacora(request: HttpRequest, pk: int):
    """
    Flujo pro:
    - Si no existe bitácora vinculada al estatus, la crea.
    - Si ya existe, la actualiza automáticamente antes de abrirla.
    """
    obj = get_object_or_404(
        EstatusOperacionalViaje.objects.select_related(
            "cliente", "conductor", "tracto", "rampla"
        ),
        pk=pk,
    )

    bitacora = (
        Bitacora.objects.filter(estatus_origen=obj)
        .order_by("-id")
        .first()
    )

    creada = False

    if not bitacora:
        bitacora = Bitacora(creado_por=request.user)
        creada = True

    bitacora = _sync_bitacora_from_estatus(
        bitacora=bitacora,
        estatus=obj,
        request_user=request.user,
        preserve_manual=True,
    )
    bitacora.save()

    if creada:
        messages.success(request, "Bitácora creada desde Estatus de Viajes.")
    else:
        messages.info(
            request,
            "La bitácora ya existía y fue actualizada automáticamente desde el estatus."
        )

    return redirect("bitacora:editar", bitacora.id)


@login_required
@permission_required("operaciones.puede_ver_estatus_viajes", raise_exception=True)
@require_GET
def estatus_viajes_planilla(request: HttpRequest):
    fecha = _parse_fecha(request.GET.get("fecha"))
    turno = (request.GET.get("turno") or "TODOS").upper()
    q = (request.GET.get("q") or "").strip()
    turnos = _turnos_a_mostrar(turno)

    rows = _qs_estatus(fecha, turnos, q)

    return render(
        request,
        "operaciones/estatus_viajes_planilla.html",
        {
            "fecha": fecha,
            "turno": turno,
            "q": q,
            "rows": rows,
        },
    )


@login_required
@permission_required("operaciones.puede_ver_estatus_viajes", raise_exception=True)
@require_GET
def estatus_viajes_export_planilla_xlsx(request: HttpRequest):
    """
    Hoy reutiliza exactamente la misma exportación del panel general.
    Si más adelante quieres un layout distinto para planilla, aquí se separa.
    """
    return estatus_viajes_export_xlsx(request)