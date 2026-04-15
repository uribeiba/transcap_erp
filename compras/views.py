from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.forms import inlineformset_factory
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from decimal import Decimal
from calendar import monthrange
from datetime import timedelta
from django.utils import timezone
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from .models import Proveedor, OrdenCompra, DetalleOrdenCompra
from .forms import ProveedorForm, OrdenCompraForm, DetalleOrdenCompraForm

# ------------------------------------------------------------
# Importación condicional de openpyxl (para evitar errores locales)
# ------------------------------------------------------------
OPENPYXL_AVAILABLE = False
Workbook = None
Font = Alignment = PatternFill = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    pass


# ========== PROVEEDORES ==========
class ProveedorListView(LoginRequiredMixin, ListView):
    model = Proveedor
    template_name = 'compras/proveedor_list.html'
    context_object_name = 'proveedores'

class ProveedorCreateView(LoginRequiredMixin, CreateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'compras/proveedor_form.html'
    success_url = reverse_lazy('compras:proveedores')
    def form_valid(self, form):
        messages.success(self.request, 'Proveedor creado correctamente.')
        return super().form_valid(form)

class ProveedorUpdateView(LoginRequiredMixin, UpdateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'compras/proveedor_form.html'
    success_url = reverse_lazy('compras:proveedores')
    def form_valid(self, form):
        messages.success(self.request, 'Proveedor actualizado.')
        return super().form_valid(form)

class ProveedorDeleteView(LoginRequiredMixin, DeleteView):
    model = Proveedor
    success_url = reverse_lazy('compras:proveedores')
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Proveedor eliminado.')
        return super().delete(request, *args, **kwargs)


# ========== ÓRDENES DE COMPRA ==========
DetalleOrdenFormSet = inlineformset_factory(
    OrdenCompra, DetalleOrdenCompra, form=DetalleOrdenCompraForm, extra=1, can_delete=True
)

class OrdenCompraListView(LoginRequiredMixin, ListView):
    model = OrdenCompra
    template_name = 'compras/orden_list.html'
    context_object_name = 'ordenes'

class OrdenCompraCreateView(LoginRequiredMixin, CreateView):
    model = OrdenCompra
    form_class = OrdenCompraForm
    template_name = 'compras/orden_form.html'
    success_url = reverse_lazy('compras:ordenes')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['detalle_formset'] = DetalleOrdenFormSet(self.request.POST, instance=self.object)
        else:
            data['detalle_formset'] = DetalleOrdenFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        detalle_formset = context['detalle_formset']
        if detalle_formset.is_valid():
            self.object = form.save(commit=False)
            self.object.creado_por = self.request.user
            self.object.save()
            detalle_formset.instance = self.object
            detalle_formset.save()
            messages.success(self.request, 'Orden de compra creada correctamente.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))

class OrdenCompraUpdateView(LoginRequiredMixin, UpdateView):
    model = OrdenCompra
    form_class = OrdenCompraForm
    template_name = 'compras/orden_form.html'
    success_url = reverse_lazy('compras:ordenes')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['detalle_formset'] = DetalleOrdenFormSet(self.request.POST, instance=self.object)
        else:
            data['detalle_formset'] = DetalleOrdenFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        detalle_formset = context['detalle_formset']
        if detalle_formset.is_valid():
            self.object = form.save()
            detalle_formset.instance = self.object
            detalle_formset.save()
            messages.success(self.request, 'Orden de compra actualizada.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))

class OrdenCompraDetailView(LoginRequiredMixin, DetailView):
    model = OrdenCompra
    template_name = 'compras/orden_detail.html'

class OrdenCompraDeleteView(LoginRequiredMixin, DeleteView):
    model = OrdenCompra
    success_url = reverse_lazy('compras:ordenes')
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Orden de compra eliminada.')
        return super().delete(request, *args, **kwargs)


# ========== DASHBOARD ==========
@login_required
def dashboard_compras(request):
    hoy = timezone.localdate()
    ordenes = OrdenCompra.objects.all()
    total_ordenes = ordenes.count()
    total_gastado = ordenes.aggregate(total=Sum('detalles__total'))['total'] or Decimal(0)
    pendientes = ordenes.filter(estado__in=['BOR', 'ENV']).count()
    primer_dia_mes = hoy.replace(day=1)
    ordenes_mes = ordenes.filter(fecha__gte=primer_dia_mes, fecha__lte=hoy)
    total_mes = ordenes_mes.aggregate(total=Sum('detalles__total'))['total'] or Decimal(0)
    cantidad_mes = ordenes_mes.count()
    top_proveedores = (OrdenCompra.objects.values('proveedor__razon_social')
                       .annotate(total=Sum('detalles__total'))
                       .order_by('-total')[:5])
    meses_labels = []
    montos_mensuales = []
    for i in range(5, -1, -1):
        fecha = hoy.replace(day=1) - timedelta(days=30*i)
        primer_dia = fecha.replace(day=1)
        ultimo_dia = primer_dia.replace(day=monthrange(primer_dia.year, primer_dia.month)[1])
        total = OrdenCompra.objects.filter(
            fecha__gte=primer_dia,
            fecha__lte=ultimo_dia
        ).aggregate(total=Sum('detalles__total'))['total'] or Decimal(0)
        meses_labels.append(primer_dia.strftime('%b %Y'))
        montos_mensuales.append(float(total))
    ultimas_ordenes = ordenes.order_by('-fecha', '-id')[:5]
    context = {
        'total_ordenes': total_ordenes,
        'total_gastado': total_gastado,
        'pendientes': pendientes,
        'total_mes': total_mes,
        'cantidad_mes': cantidad_mes,
        'top_proveedores': top_proveedores,
        'meses_labels': meses_labels,
        'montos_mensuales': montos_mensuales,
        'ultimas_ordenes': ultimas_ordenes,
    }
    return render(request, 'compras/dashboard.html', context)


# ========== EXPORTACIONES ==========
@login_required
def exportar_ordenes_excel(request):
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(
            "Exportación a Excel no disponible. Instale openpyxl: pip install openpyxl", 
            status=503
        )
    ordenes = OrdenCompra.objects.select_related('proveedor').order_by('-fecha', '-id')
    wb = Workbook()
    ws = wb.active
    ws.title = "Órdenes de Compra"
    headers = ['Nº Orden', 'Fecha', 'Proveedor', 'RUT Proveedor', 'Total', 'Estado', 'Observaciones']
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for orden in ordenes:
        ws.append([
            orden.numero,
            orden.fecha.strftime('%d/%m/%Y'),
            orden.proveedor.razon_social,
            orden.proveedor.rut,
            float(orden.total()),
            orden.get_estado_display(),
            orden.observaciones[:100] if orden.observaciones else ''
        ])
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[col_letter].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ordenes_compra.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_ordenes_pdf(request):
    """Genera PDF de órdenes de compra usando ReportLab"""
    ordenes = OrdenCompra.objects.select_related('proveedor').order_by('-fecha', '-id')
    
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    x = 20 * mm
    y = height - 20 * mm
    
    # Título
    p.setFont("Helvetica-Bold", 14)
    p.drawString(x, y, "Transcap ERP - Órdenes de Compra")
    y -= 10 * mm
    
    # Fecha de generación
    p.setFont("Helvetica", 9)
    p.drawString(x, y, f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    y -= 10 * mm
    
    # Encabezados de tabla
    p.setFont("Helvetica-Bold", 8)
    p.drawString(x, y, "Nº Orden")
    p.drawString(x + 30 * mm, y, "Fecha")
    p.drawString(x + 55 * mm, y, "Proveedor")
    p.drawString(x + 110 * mm, y, "Total")
    p.drawString(x + 140 * mm, y, "Estado")
    y -= 5 * mm
    p.line(x, y, x + 170 * mm, y)
    y -= 4 * mm
    
    p.setFont("Helvetica", 8)
    for orden in ordenes:
        if y < 40 * mm:
            p.showPage()
            y = height - 20 * mm
            p.setFont("Helvetica", 8)
        
        p.drawString(x, y, orden.numero)
        p.drawString(x + 30 * mm, y, orden.fecha.strftime('%d/%m/%Y'))
        p.drawString(x + 55 * mm, y, orden.proveedor.razon_social[:35])
        p.drawString(x + 110 * mm, y, f"${orden.total():,.0f}")
        p.drawString(x + 140 * mm, y, orden.get_estado_display())
        y -= 5 * mm
    
    # Pie de página
    y = 20 * mm
    p.setFont("Helvetica-Oblique", 7)
    p.drawString(x, y, "Documento generado por Transcap ERP - www.transcap.cl")
    
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ordenes_compra.pdf"'
    response.write(buffer.getvalue())
    return response